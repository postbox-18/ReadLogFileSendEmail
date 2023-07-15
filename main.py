import smtplib

import threading
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import schedule

from flask_restful import Resource, Api
import openpyxl
from openpyxl.reader.excel import load_workbook
from email.mime.application import MIMEApplication
from flask import Flask
import os.path

import os
import time
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler

app = Flask("HomeAPI")
api = Api(app)


# LOG FILE FOR PRINT MESSAGE INITIALLY REMOVE
def LogFileRemove():
    print("LogFileRemove")
    if os.path.exists("log.txt"):
        os.remove("log.txt")
    else:
        pass


# LOG FILE FOR PRINT MESSAGE CREATE AND WRITE
def text(filename):
    s1 = filename
    f1 = open("log.txt", "a")
    f1.write(str(s1) + " \n")
    f1.close()


# SSL FOR SEND EMAIL
def SSL(html, log_file, attach):
    try:
        # SENDER
        text("SENDER")
        # READ LOG DETAILS FILE 
        filename = "Log Details.xlsx"
        LogDetailsXl = openpyxl.load_workbook(filename)
        LogDetailsXlSheet1 = LogDetailsXl["Details"]
        sender_email = LogDetailsXlSheet1['B1'].value.strip()
        password = LogDetailsXlSheet1['B2'].value.strip()
        emailSubject = LogDetailsXlSheet1['B3'].value.strip()
        port = LogDetailsXlSheet1['B4'].value
        smtp_server = LogDetailsXlSheet1['B5'].value.strip()

        # RECEIVER LIST
        text("Receiver List")
        receiverSheet = LogDetailsXl["TO"]
        # COUNT NUMBER OF ROWS
        n = receiverSheet.max_row + 1

        # GET NUMBER RECEIVER LIST
        text(n)
        input = []
        for i in range(1, n):
            input.append(receiverSheet['A' + str(i)].value.strip())
        receiver_email = ", ".join(input)
        text(receiver_email)

        # SET SSL 
        message = MIMEMultipart("alternative")
        message["Subject"] = emailSubject + " Log File " + log_file.strip()
        message["From"] = sender_email
        message["To"] = receiver_email

        # CHECK IF HTML IS EMPTY ATTACH PDF ZIP XLSX FILE NOT TABLE FORMAT TEXT
        if not html:
            attach = MIMEApplication(open(attach, 'rb').read())
            attach.add_header('Content-Disposition', 'attachment', filename=attach)
            message.attach(attach)
        else:
            htmlFormat = MIMEText(html, "html")
            message.attach(htmlFormat)

        smtp_server = smtplib.SMTP_SSL(smtp_server, port)
        smtp_server.ehlo()
        smtp_server.login(sender_email, password)
        smtp_server.sendmail(sender_email, input, message.as_string())
        smtp_server.close()

        # AFTER SENDING EMAIL ATTACHMENT DELETE XLSX SHEET
        if html == "":
            os.remove(attach)
    except Exception as e:
        text(e)
        


# CREATE TABLE FORMAT DATA
def CreateTable(headerList, Data, log_file):
    text("CreateTable")
    # APPEND HEADER IN TABLE LIST
    table_html = '<html><table border="2">\n'
    table_html += '<tr>\n'
    for value in headerList:
        table_html += f'<th>{value}</th>\n'
    table_html += '</tr>\n'

    # APPEND VALUE IN TABLE LIST
    for row in Data:
        table_html += '<tr>\n'
        for value in row:
            table_html += f'<td>{value}</td>\n'
        table_html += '</tr>\n'
    table_html += '</table></html>'

    SSL(table_html, log_file, "")


def ReadLog(log_file):
    try:
        # READ LOG READ FILE AND STORE TO XLSX FILE
        logReadList = "logReadList.xlsx"
        # ADD ATTACHMENT FILE AT END OF DAY TO SEND
        attach = "attachment.xlsx"
        # DETAILS IN XLSX FILE
        filename = "Log Details.xlsx"

        # CREATE XLSX AND ADD DATA
        with open(log_file, 'r') as file:
            log_data = file.readlines()

        xlsx_data = []
        headerList = None
        # SPLIT THE LOG DATA IN LIST
        for line in log_data:
            if line.startswith("#Fields:"):
                headerList = line[9:].strip().split()
            elif not line.startswith("#"):
                fields = line.strip().split()
                xlsx_data.append(fields)

        # CREATE LOG READ XLSX FILE
        create_xlsx = openpyxl.Workbook()
        create_xlsx.save(logReadList)

        # READ LOG READ XLSX SHEET
        logReadXl = load_workbook(logReadList)
        logReadXlSheet = logReadXl.worksheets[0]
        del logReadXl[logReadXl.active.title]
        logReadXl.create_sheet()
        logReadXlSheet = logReadXl.worksheets[0]

        # ADD VALUE IN LOG READ XLSX
        logReadXlSheet.append(headerList)
        for value in xlsx_data:
            logReadXlSheet.append(value)
        logReadXl.save(logReadList)
        text('Log file converted to XLSX successfully.')

        # CHECK CONDITION IN LOG READ XLSX
        logReadXlSheetCheck = logReadXl.worksheets[0]
        max_row = logReadXlSheetCheck.max_row + 1
        max_column = logReadXlSheetCheck.max_column + 1
        find_status = 0
        for i in range(1, max_column):
            if logReadXlSheetCheck.cell(column=i, row=1).value.lower() == "sc-status":
                find_status = i
                break
        text(find_status)
        # ADD DATA IN ATTACH XLSX
        attach_Data = []
        critical_Data = []
        
        # READ LOG DETAILS FILE 
        LogDetailsXl = openpyxl.load_workbook(filename)
        LogDetailsXlSheet1 = LogDetailsXl["Details"]
        status = str(LogDetailsXlSheet1['B7'].value).strip()
        criticalStatus = str(LogDetailsXlSheet1['B8'].value).strip()
        
        for i in range(2, max_row):
            if logReadXlSheet.cell(column=find_status, row=i).value != status:
                # CHECK IF STATUS IS CRITICAL OR NOT
                if logReadXlSheet.cell(column=find_status, row=i).value == criticalStatus:
                    column_data = []
                    for j in range(1, max_column):
                        column_data.append(logReadXlSheet.cell(column=j, row=i).value)
                    # ADD LOG PATH TO COLUMN
                    column_data.append(log_file)
                    # ADD BOTH LIST FOR SEND MAIL AND END OF DAY STATEMENT 
                    critical_Data.append(column_data)
                    attach_Data.append(column_data)


                else:
                    column_data = []
                    for j in range(1, max_column):
                        column_data.append(logReadXlSheet.cell(column=j, row=i).value)
                    # ADD LOG PATH TO COLUMN
                    column_data.append(log_file)
                    attach_Data.append(column_data)
        # ADD BLANK ROW
        attach_Data.append([""])

        text(attach_Data)
        text(critical_Data)
        # CREATE ATTACH XLSX SHEET
        try:
            readattach = load_workbook(attach)
        except Exception as e:
            text("e=>" + str(e))
            # CREATE ATTACH XLSX
            create_xlsx = openpyxl.Workbook()
            sheets = create_xlsx.worksheets[0]
            headerList.append("Log File Path")
            sheets.append(headerList)
            create_xlsx.save(attach)
            readattach = load_workbook(attach)
        readattach.save(attach)

        # ADD ATTACH XLSX SHEET
        attachsheet1 = readattach.worksheets[0]
        for value in attach_Data:
            attachsheet1.append(value)
        readattach.save(attach)

        # SEND EMAIL WHEN CRITICAL DATA
        if critical_Data:
            CreateTable(headerList, critical_Data, log_file)

    except Exception as e:
        text(e)
        


def TimeStamp(attach):
    LogFileRemove()

    print("TimeStamp")
    text("TimeStamp")
    try:
        # SEND EMAIL WHEN END OF DAY
        def testSSL():
            SSL("", "", attach)

        # READ LOG DETAILS FILE
        LogDetailsXl = openpyxl.load_workbook('Log Details.xlsx')
        LogDetailsXlSheet = LogDetailsXl["Details"]
        timeStamp = LogDetailsXlSheet['B9'].value.strip()
        schedule.every().day.at(timeStamp).do(testSSL)

        while True:
            schedule.run_pending()
            time.sleep(1)
    except Exception as e:
        text(e)
        


class MonitorFolder(FileSystemEventHandler):
    FILE_SIZE = 1000

    def on_created(self, event):
        LogFileRemove()
        text('Create')
        text(event.src_path + "\t" + event.event_type)
        ReadLog(event.src_path)


def Monitor():
    try:
        # CLEAR LOG FILE
        LogFileRemove()
        # MONITOR FOLDER
        filename = "Log Details.xlsx"
        LogDetailsXl = openpyxl.load_workbook(filename)
        LogDetailsXlSheet = LogDetailsXl["Details"]
        src_path = LogDetailsXlSheet['B6'].value.strip()
        text(src_path)
        event_handler = MonitorFolder()
        observer = Observer()
        observer.schedule(event_handler, path=src_path, recursive=True)
        text("Monitoring started")

        observer.start()
        try:
            while (True):
                time.sleep(1)

        except Exception as e:

            text(e)
            
            observer.stop()
            observer.join()
    except Exception as e:
        text(e)
        


class Home(Resource):
    def get(self):
        text("Completed")
        # TIME STAMP
        threading.Thread(target=TimeStamp, args=("attachment.xlsx",)).start()
        threading.Thread(target=Monitor).start()

        return "completed"


api.add_resource(Home, '/')

if __name__ == '__main__':
    while True:
        app.run()
